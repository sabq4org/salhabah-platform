#!/usr/bin/env python3
"""
Extract teachers + administrators from a school-supplied Excel sheet.

Source format (the school's "بيانات المعلمات والاداريات" workbook):
  Section 1 (rows 5..N):  المعلمات
      headers: الاسم | السجل المدني | تاريخ الميلاد | رقم الجوال | البريد الوزاري

  Section 2 (rows N+2..M): الاداريات والمستخدمين
      headers: الاسم | السجل المدني | المسمى الوظيفي | العمل الحالي | تاريخ المباشرة | رقم الجوال

The two sections live in a SINGLE worksheet, separated by a blank row plus a
section-header row (e.g. "بيانات المعلمات" / "بيانات الاداريات والمستخدمين").

Output:
  scripts/import-data/teachers-extra.json   (rows for db:import-teachers)
  scripts/import-data/admins.json           (rows for db:import-admins)

Phone-number normalisation:
  - 9 digits starting with '5'  → '05XXXXXXXX'  (Saudi mobile)
  - 10 digits starting with '5' → take first 9 then prefix with '0'
                                  (handles the data-entry mistake we saw)
  - 10 digits starting with '0' → kept as-is
  - blank / unparseable         → null

Hijri DOB cells are kept verbatim and stored in `notes` since the schema has
no separate hijri_dob column.
"""

from __future__ import annotations
import json
import re
import sys
from pathlib import Path

import openpyxl

SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/teachers-and-admins.xlsx")
OUT_DIR = Path(__file__).parent / "import-data"
OUT_TEACHERS = OUT_DIR / "teachers-extra.json"
OUT_ADMINS = OUT_DIR / "admins.json"


def norm_phone(raw) -> str | None:
    """Normalise a Saudi mobile number to '05XXXXXXXX' or return None."""
    if raw is None:
        return None
    digits = re.sub(r"\D+", "", str(raw))
    if not digits:
        return None
    if len(digits) == 10 and digits.startswith("05"):
        return digits
    if len(digits) == 9 and digits.startswith("5"):
        return "0" + digits
    if len(digits) == 10 and digits.startswith("5"):
        # 10 digits without leading 0 but starting with 5 — assume the trailing
        # digit was an accidental paste (the school sheet had one such row),
        # take the first 9 digits and prefix 0.
        return "0" + digits[:9]
    if len(digits) == 9 and digits.startswith("05"):
        return digits
    # Anything else: keep verbatim so it can be reviewed later
    return digits


def department_for(job_title: str, current_role: str) -> str:
    """Best-effort mapping from 'العمل الحالي' to a department string."""
    blob = f"{current_role or ''} {job_title or ''}"
    if "أمن" in blob or "امن" in blob or "سلامة" in blob:
        return "الأمن والسلامة"
    if "نور" in blob:
        return "نظام نور"
    if "حارس" in blob or "حراسة" in blob:
        return "الحراسة"
    if "نشاط" in blob:
        return "النشاط"
    if "مالي" in blob or "محاسب" in blob:
        return "الشؤون المالية"
    return "الشؤون الإدارية"


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[list] = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # Locate the two section header rows
    teacher_header_row = None
    admin_header_row = None
    for i, row in enumerate(rows):
        joined = " ".join(str(c) for c in row if c)
        if "الاسم" in joined and "البريد الوزاري" in joined:
            teacher_header_row = i
        elif "الاسم" in joined and "المسمى" in joined:
            admin_header_row = i

    if teacher_header_row is None or admin_header_row is None:
        raise SystemExit(f"Could not locate headers (teachers={teacher_header_row}, admins={admin_header_row})")

    # Teachers: rows after teacher_header_row up to (but not including) the
    # section-divider that precedes the admin header.
    teachers: list[dict] = []
    for i in range(teacher_header_row + 1, admin_header_row):
        row = rows[i]
        full_name = (row[0] or "").strip() if row[0] else ""
        national_id = str(row[1]).strip() if row[1] else ""
        if not full_name or not national_id:
            continue
        if not re.fullmatch(r"\d{8,15}", national_id):
            continue
        dob_hijri = (str(row[2]).strip() if row[2] else "") or None
        phone = norm_phone(row[3])
        email = (str(row[4]).strip() if row[4] else "") or None
        notes_bits = []
        if dob_hijri:
            notes_bits.append(f"تاريخ الميلاد (هـ): {dob_hijri}")
        teachers.append({
            "fullName": full_name,
            "nationalId": national_id,
            "phone": phone,
            "email": email,
            "notes": " | ".join(notes_bits) if notes_bits else None,
        })

    # Admins: rows after admin_header_row to end of sheet.
    admins: list[dict] = []
    for i in range(admin_header_row + 1, len(rows)):
        row = rows[i]
        full_name = (row[0] or "").strip() if row[0] else ""
        national_id = str(row[1]).strip() if row[1] else ""
        if not full_name or not national_id:
            continue
        if not re.fullmatch(r"\d{8,15}", national_id):
            continue
        job_title = (str(row[2]).strip() if row[2] else "") or "مساعد اداري"
        current_role = (str(row[3]).strip() if row[3] else "") or job_title
        hire_date_hijri = (str(row[4]).strip() if row[4] else "") or None
        phone = norm_phone(row[5])
        notes_bits = []
        if current_role and current_role != job_title:
            notes_bits.append(f"العمل الحالي: {current_role}")
        if hire_date_hijri:
            notes_bits.append(f"تاريخ المباشرة (هـ): {hire_date_hijri}")
        # flag the suspicious phone case so the admin can fix it later
        raw_phone_str = re.sub(r"\D+", "", str(row[5] or ""))
        if len(raw_phone_str) == 10 and raw_phone_str.startswith("5"):
            notes_bits.append(f"⚠️ رقم الجوال الأصلي: {raw_phone_str} — يحتاج تأكيد")
        admins.append({
            "fullName": full_name,
            "nationalId": national_id,
            "jobTitle": job_title,
            "department": department_for(job_title, current_role),
            "phone": phone,
            "notes": " | ".join(notes_bits) if notes_bits else None,
        })

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_TEACHERS.write_text(json.dumps(teachers, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_ADMINS.write_text(json.dumps(admins, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"✓ Wrote {len(teachers)} teachers → {OUT_TEACHERS}")
    print(f"✓ Wrote {len(admins)} admins   → {OUT_ADMINS}")
    print()
    print("Teachers preview:")
    for t in teachers:
        print(f"  - {t['nationalId']}  {t['fullName']}  | {t['email'] or '-'}")
    print()
    print("Admins preview:")
    for a in admins:
        print(f"  - {a['nationalId']}  {a['fullName']}  | {a['jobTitle']} | {a['department']} | {a['phone'] or '-'}")


if __name__ == "__main__":
    main()
